import settings

from database import qtd_invoices, last_send_mails, last_invoice
from openpyxl import Workbook


class ReportCreationException(Exception):
    """ Exceção lançada ao criar a planilha de requisições de envio de faturas"""


def make_invoices_report(invoice_report):
    try:
        # Criação dos dados do relatório
        wb = Workbook()

        # Folha de envio de emails
        send_mails_sheet = wb.active
        send_mails_sheet.title = "Demora no envio do e-mail"

        for row in last_send_mails():
            send_mails_sheet.append(row)

        queue_invoices_sheet = wb.create_sheet(title="Enfileiramento de requisicao")
        for row in qtd_invoices():
            queue_invoices_sheet.append(row)

        invoices_latence_sheet = wb.create_sheet(title="Latencia na requisicao")

        for row in last_invoice():
            invoices_latence_sheet.append(row)

        wb.save(filename=invoice_report)

        return True
    except Exception:
        return False


def main():
    print("Criando o relatório...")
    try:
        if make_invoices_report(settings.INVOICE_REPORT_PATH):
            print('Relatório criado com sucesso em %s' % settings.INVOICE_REPORT_PATH)
        else:
            raise ReportCreationException('Falha ao criar o relatório')
    except ReportCreationException as report_exc:
        print(report_exc)


if __name__ == '__main__':
    main()
